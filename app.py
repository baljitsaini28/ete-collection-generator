from __future__ import annotations

import math
import os
import tempfile
import uuid
from collections import defaultdict
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_TITLE = "ETE Collection Format Generator"
PORT = int(os.environ.get("PORT", "8765"))
HOST = os.environ.get("HOST", "0.0.0.0" if "PORT" in os.environ else "127.0.0.1")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)


SUBJECTIVE_TYPES = {"all subjective", "mix mcq + subjective"}


def norm(value) -> str:
    return str(value or "").strip()


def as_count(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def is_subjective_type(value) -> bool:
    return norm(value).lower() in SUBJECTIVE_TYPES


def is_omr_type(value) -> bool:
    text = norm(value).lower()
    return text == "mix mcq + subjective" or "objective" in text


def parse_grid(grid_path: Path):
    wb = load_workbook(grid_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [norm(ws.cell(1, col).value).lower() for col in range(1, ws.max_column + 1)]
    try:
        course_col = headers.index("course code") + 1
        type_col = headers.index("paper type") + 1
    except ValueError as exc:
        raise ValueError("Grid must contain 'Course Code' and 'Paper Type' headers in row 1.") from exc

    sum_col = headers.index("sum") + 1 if "sum" in headers else ws.max_column
    room_cols = [
        (col, norm(ws.cell(1, col).value))
        for col in range(type_col + 2, sum_col)
        if norm(ws.cell(1, col).value)
    ]

    subjective = defaultdict(lambda: {"total": 0, "subpacket": 0})
    omr = defaultdict(int)

    for row in range(2, ws.max_row + 1):
        course = norm(ws.cell(row, course_col).value)
        paper_type = ws.cell(row, type_col).value
        if not course:
            continue

        total = as_count(ws.cell(row, sum_col).value)

        if is_subjective_type(paper_type):
            subjective[course]["total"] += total
            subjective[course]["subpacket"] += sum(
                1 for col, _room in room_cols if as_count(ws.cell(row, col).value) > 0
            )

        if is_omr_type(paper_type):
            for col, room in room_cols:
                count = as_count(ws.cell(row, col).value)
                if count > 0:
                    omr[(room, course)] += count

    return subjective, omr


def sheet_styles():
    thin = Side(style="thin", color="000000")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_fill": PatternFill("solid", fgColor="D9EAF7"),
        "total_fill": PatternFill("solid", fgColor="FFF2CC"),
    }


def style_used_range(ws):
    styles = sheet_styles()
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = styles["border"]
            if cell.row in (1, 2, 4):
                cell.font = Font(bold=True)
            if cell.row == 4:
                cell.fill = styles["header_fill"]


def style_total_row(ws, row, max_col):
    styles = sheet_styles()
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = styles["total_fill"]
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_common_top(ws, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A1"] = "Lovely Professional University, Phagwara"
    ws["A2"] = "Collection Center Details"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22


def add_side_details(ws, start_col):
    details = [
        (4, "Write UID and Phone number of the Person who has prepared the data here: Dr. Baljit Singh Saini"),
        (6, "Center No. ", 2509),
        (8, "Block/Room No.:", "36-802A"),
        (10, "Date:    ", "25-05-2026"),
        (12, "Session(Time): ", "S2"),
        (14, "email this filled performa just after every session to UID's: 14623 and 20506 through LPU internal email."),
    ]
    for row, label, *value in details:
        ws.cell(row, start_col, label)
        if value:
            ws.cell(row, start_col + 1, value[0])


def prepare_sheet(wb, name, max_col):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    for row in range(5, max(ws.max_row, 5) + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None
    return ws


def create_or_load_template(template_path: Path | None):
    if template_path and template_path.suffix.lower() == ".xlsx":
        try:
            return load_workbook(template_path)
        except Exception:
            pass
    wb = Workbook()
    wb.active.title = "Subjective Sheets"
    wb.create_sheet("OMR Sheets")
    return wb


def fill_subjective(ws, subjective):
    set_common_top(ws, 8)
    headers = ["S.No", "Course Code", "Total Sheets", "Packet", "SubPacket", "UMC", "", ""]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    add_side_details(ws, 7)

    for idx, course in enumerate(sorted(subjective), start=1):
        total = subjective[course]["total"]
        row = idx + 4
        values = [idx, course, total, math.ceil(total / 100) if total else 0, subjective[course]["subpacket"], ""]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value)

    total_row = len(subjective) + 5
    ws.cell(total_row, 2, "Total")
    for col in (3, 4, 5):
        ws.cell(total_row, col, f"=SUM({ws.cell(5, col).coordinate}:{ws.cell(total_row - 1, col).coordinate})")

    for col, width in enumerate([8, 18, 14, 10, 12, 10, 55, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    style_used_range(ws)
    style_total_row(ws, total_row, 8)


def fill_omr(ws, omr):
    set_common_top(ws, 7)
    headers = [
        "Center No.",
        "Room No.",
        "Course Code",
        "Total OMR Sheets",
        "UMC",
        "Write UID and Phone number of the Person who has prepared the data here: ",
        "",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(4, col, value)
    add_side_details(ws, 6)

    for idx, ((room, course), count) in enumerate(sorted(omr.items(), key=lambda item: (item[0][0], item[0][1])), start=1):
        for col, value in enumerate([2509, room, course, count, ""], start=1):
            ws.cell(idx + 4, col, value)

    total_row = len(omr) + 5
    ws.cell(total_row, 3, "Total")
    ws.cell(total_row, 4, f"=SUM(D5:{ws.cell(total_row - 1, 4).coordinate})")

    for col, width in enumerate([12, 14, 18, 18, 10, 55, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    style_used_range(ws)
    style_total_row(ws, total_row, 7)


def build_output(grid_path: Path, template_path: Path | None) -> Path:
    subjective, omr = parse_grid(grid_path)
    wb = create_or_load_template(template_path)
    fill_subjective(prepare_sheet(wb, "Subjective Sheets", 8), subjective)
    fill_omr(prepare_sheet(wb, "OMR Sheets", 7), omr)

    output_path = OUTPUT_DIR / f"ETE_Collection_Filled_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(output_path)
    return output_path


HTML = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>ETE Collection Format Generator</title>
  <style>
    :root {
      color-scheme: light;
      --ink: #1e293b;
      --muted: #64748b;
      --line: #d9e2ec;
      --paper: #ffffff;
      --soft: #f3f7fb;
      --accent: #0f766e;
      --accent-dark: #115e59;
      --gold: #b7791f;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      font-family: Inter, Segoe UI, Arial, sans-serif;
      color: var(--ink);
      background:
        linear-gradient(135deg, rgba(15, 118, 110, .10), rgba(183, 121, 31, .10)),
        var(--soft);
      display: grid;
      place-items: center;
      padding: 32px;
    }
    main {
      width: min(920px, 100%);
      background: var(--paper);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 24px 70px rgba(30, 41, 59, .14);
      overflow: hidden;
    }
    header {
      padding: 30px 34px 24px;
      border-bottom: 1px solid var(--line);
      background: #fbfdff;
    }
    h1 {
      margin: 0 0 8px;
      font-size: 30px;
      line-height: 1.15;
      letter-spacing: 0;
    }
    p { margin: 0; color: var(--muted); line-height: 1.55; }
    form { padding: 30px 34px 34px; display: grid; gap: 22px; }
    .grid { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 18px; }
    label {
      display: grid;
      gap: 10px;
      font-weight: 700;
      color: #334155;
    }
    .hint {
      display: block;
      font-weight: 500;
      color: var(--muted);
      font-size: 13px;
    }
    input[type=file] {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 13px;
      background: #fff;
      color: var(--ink);
    }
    .panel {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      background: #fbfdff;
      display: grid;
      gap: 8px;
    }
    .checks {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 10px;
      color: #475569;
      font-size: 14px;
    }
    .check {
      border: 1px solid #e5edf5;
      border-radius: 8px;
      padding: 10px 12px;
      background: #fff;
    }
    button {
      justify-self: start;
      border: 0;
      border-radius: 8px;
      padding: 13px 20px;
      background: var(--accent);
      color: white;
      font-weight: 800;
      cursor: pointer;
      box-shadow: 0 10px 24px rgba(15, 118, 110, .24);
    }
    button:hover { background: var(--accent-dark); }
    footer {
      padding: 16px 34px;
      border-top: 1px solid var(--line);
      color: var(--muted);
      font-size: 13px;
      background: #fbfdff;
    }
    @media (max-width: 760px) {
      body { padding: 14px; place-items: start center; }
      header, form, footer { padding-left: 20px; padding-right: 20px; }
      .grid, .checks { grid-template-columns: 1fr; }
      h1 { font-size: 24px; }
      button { width: 100%; }
    }
  </style>
</head>
<body>
  <main>
    <header>
      <h1>ETE Collection Format Generator</h1>
      <p>Upload the room-wise exam grid and, optionally, the ETE collection format. The generated Excel file includes Subjective Sheets, OMR Sheets, packets, subpackets, and totals.</p>
    </header>
    <form action="/convert" method="post" enctype="multipart/form-data">
      <div class="grid">
        <label>
          Grid file
          <span class="hint">Required .xlsx file containing Course Code, Paper Type, room columns, and Sum.</span>
          <input type="file" name="grid" accept=".xlsx" required>
        </label>
        <label>
          ETE collection format
          <span class="hint">Optional .xls or .xlsx template. The app uses the fixed format if this is not selected.</span>
          <input type="file" name="template" accept=".xls,.xlsx">
        </label>
      </div>
      <section class="panel" aria-label="Generated output includes">
        <div class="checks">
          <div class="check">Subjective course summary</div>
          <div class="check">Room-wise OMR entries</div>
          <div class="check">Bottom total rows</div>
        </div>
      </section>
      <button type="submit">Start Conversion</button>
    </form>
    <footer>Runs locally on this computer. Uploaded files are used only to produce the downloadable Excel output.</footer>
  </main>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def parse_uploads(self):
        content_type = self.headers.get("Content-Type", "")
        content_length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(content_length)
        message = BytesParser(policy=policy.default).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
        )
        uploads = {}
        if not message.is_multipart():
            return uploads
        for part in message.iter_parts():
            disposition = part.get("Content-Disposition", "")
            if "form-data" not in disposition:
                continue
            name = part.get_param("name", header="content-disposition")
            filename = part.get_filename()
            if not name:
                continue
            uploads[name] = {
                "filename": filename or "",
                "data": part.get_payload(decode=True) or b"",
            }
        return uploads

    def do_GET(self):
        if self.path == "/" or self.path == "/index.html":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(HTML.encode("utf-8"))
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self):
        if self.path != "/convert":
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        try:
            uploads = self.parse_uploads()
            grid_item = uploads.get("grid")
            template_item = uploads.get("template")
            if grid_item is None or not grid_item["filename"]:
                raise ValueError("Please select the grid .xlsx file.")

            with tempfile.TemporaryDirectory() as tmp:
                tmp_path = Path(tmp)
                grid_path = tmp_path / Path(grid_item["filename"]).name
                grid_path.write_bytes(grid_item["data"])

                template_path = None
                if template_item is not None and template_item["filename"]:
                    template_path = tmp_path / Path(template_item["filename"]).name
                    template_path.write_bytes(template_item["data"])

                output_path = build_output(grid_path, template_path)

            data = output_path.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="ETE Collection Filled.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as exc:
            message = f"Conversion failed: {exc}"
            self.send_response(HTTPStatus.BAD_REQUEST)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write(message.encode("utf-8"))


def main():
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"{APP_TITLE} running at http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
